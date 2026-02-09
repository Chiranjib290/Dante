from openpyxl import Workbook
import requests


BASE_URL="https://dpe.pwc.com"
TIMEOUT = 30
AUTH = ("chiranjib.bhattacharyya@in.pwc.com","Change@123456")
XFs=set()
RESULT=set()

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)



url = f"{BASE_URL}/bin/querybuilder.json?1_property=jcr%3acontent%2froot%2frightrailcontact%2fsling%3aresourceType&1_property.value=pwc%2fcomponents%2fmodernized%2fcontent%2frightrailcontact&2_property=jcr%3acontent%2froot%2frightrailcontact%2flinks&2_property.operation=exists&2_property.value=false&p.hits=selective&p.limit=-1&p.properties=jcr%3apath%20jcr%3acontent%2froot%2frightrailcontact%2flinks%20jcr%3acontent%2froot%2frightrailcontact%2fsling%3aresourceType&path=%2fcontent%2fexperience-fragments%2fpwc%2fuk%2fen%2fcontact%2f"
try:
        resp = requests.get(url, auth=(AUTH), timeout=TIMEOUT)
except Exception as e:
    print(f"Failed : {e}")

if resp.status_code != 200:
    print(f"Unexpected HTTP {resp.status_code} when fetching")

try:
    payload = resp.json()
except ValueError:
    print(f"Invalid JSON returned")

hits = payload.get("hits")
if hits ==[]:
    print(f"No hits")
else:
    for hit in hits:
        path=hit.get("jcr:path")
        XFs.add(path)
        print(f"XF path - {path}")
        uri=f"{BASE_URL}/bin/wcm/references.json?path={path}"
        try:
            respi = requests.get(uri, auth=(AUTH), timeout=TIMEOUT)
        except Exception as e:
            print(f"Failed : {e}")

        if respi.status_code != 200:
            print(f"Unexpected HTTP {respi.status_code} when fetching")

        try:
            payloadi = respi.json()
        except ValueError:
            print(f"Invalid JSON returned")

        pages = payloadi.get("pages")
        if pages ==[]:
            print(f"No Pages")
        else:
            for page in pages:
                #isPub = page.get("published")
                #isPage = page.get("isPage")
                refs = page.get("references")
                for ref in refs:
                    if ref.startswith("/content/pwc/uk/en/"):
                        ref = ref.split("/jcr:content/")[0]
                        # ref = ref.replace("/content/pwc/uk/en/","https://www.pwc.co.uk/")
                        # ref = ref + ".html"
                        ref = f"{BASE_URL}/editor.html" + ref + ".html"
                        #if requests.get(ref).status_code==200: 
                        RESULT.add(ref)


ls = list(RESULT)
print(ls)
file="empty_contact_us_s.xlsx"
list_to_excel(list(RESULT),file)