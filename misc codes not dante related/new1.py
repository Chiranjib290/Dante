import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────

INPUT_FILE = "test1102.xlsx"
OUTPUT_FILE = INPUT_FILE  # overwrite in place

BASE_URL    = 'http://10.195.129.69:4503'
BASE_URL_2  = 'http://10.195.129.68:4503'
AUTH        = ('admin', 'BSpr0d@A3m2019')

MAX_WORKERS = 50

# ──────────────────────────────────────────────────────────────────────────────
# GLOBALS
# ──────────────────────────────────────────────────────────────────────────────

wb   = load_workbook(INPUT_FILE)
ws   = wb.active
lock = threading.Lock()

# Thread‐local storage for per‐thread Session
thread_local = threading.local()

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def get_session():
    """
    Return a thread‐local Session, creating it on first use.
    """
    if not hasattr(thread_local, "session"):
        sess = requests.Session()
        sess.auth = AUTH
        thread_local.session = sess
    return thread_local.session

def user_delete(row_idx, path):
    """
    Attempt to delete a user at `path` on both servers.
    Write status strings back to columns B and C.
    """
    sess = get_session()
    out1 = out2 = ""

    # First server
    try:
        r = sess.get(f"{BASE_URL}{path}.json", timeout=20)
        if r.status_code == 200:
            p = sess.post(f"{BASE_URL}{path}",
                          data={"deleteAuthorizable": ""},
                          timeout=20)
            out1 = "Deleted" if p.status_code == 200 else f"Failed to delete"
        else:
            out1 = f"Unable to find user"
    except Exception as e:
        out1 = f"Error: {e}"

    # Second server
    try:
        r2 = sess.get(f"{BASE_URL_2}{path}.json", timeout=20)
        if r2.status_code == 200:
            p2 = sess.post(f"{BASE_URL_2}{path}",
                           data={"deleteAuthorizable": ""},
                           timeout=20)
            out2 = "Deleted" if p2.status_code == 200 else f"Failed to delete"
        else:
            out2 = f"Unable to find user"
    except Exception as e:
        out2 = f"Error: {e}"

    # Write back to the sheet under thread‐safety lock
    with lock:
        ws.cell(row=row_idx, column=2, value=out1)
        ws.cell(row=row_idx, column=3, value=out2)

def main():
    # Submit one task per row (skip header in column A)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as exe:
        futures = [
            exe.submit(user_delete, idx, cell.value)
            for idx, cell in enumerate(ws['A'][1:], start=2)
        ]

        # Propagate exceptions if any
        for f in as_completed(futures):
            f.result()

    # Save the updated workbook
    wb.save(OUTPUT_FILE)

if __name__ == "__main__":
    main()
