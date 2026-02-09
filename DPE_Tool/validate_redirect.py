import logging
import json
from time import sleep
import openpyxl
import requests

class RedirectValidation:
    headers = {
                "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                "accept-encoding": "gzip, deflate, br",
                "accept-language": "en-US,en;q=0.9",
                "cache-control": "max-age=0",
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
            }
    http_response = {
        100: "Continue",
        101: "Switching Protocols",
        200: "Ok",
        201: "Created",
        202: "Accepted",
        203: "Non-Authoritative Information",
        204: "Partial Conte",
        301: "Moved Permanently",
        302: "Found",
        307: "Temporary Redirect",
        400: "Bad Reques",
        401: "Unauthorised",
        403: "Forbidden",
        404: "Page not found",
        405: "Method not allowed",
        406: "Not acceptable",
        408: "Request timeout",
        410: "Gone",
        413: "Payload too large",
        414: "URI Too Long",
        429: "Too Many request",
        500: "Internal server error",
        502: "Bad Gateway",
        503: "Service Unavailable",
        504: "Gateway timeout",
        999: "Exception Occurred"
    }
    def __init__(self):
        log_level = {
            'debug': logging.DEBUG,
            'info': logging.INFO,
            'warning': logging.WARNING,
            'error': logging.ERROR,
            'critical': logging.CRITICAL
        }
        self.configdata = self.opencode("configfiles\\config.json")
        self.loglevel = self.configdata["loglevel"]
        self.sleeptime = self.configdata["sleeptime"]
        self.timeout = self.configdata["timeout"]
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level[self.loglevel])
        self.logger.debug("Loglevel: "+str(self.loglevel)+", Sleeptime: "+str(self.sleeptime)+", Timeout: "+str(self.timeout))

    def remove_trailing_slashes(self, url):
        try:
            url = str(url)
            if url.strip() != "":
                while (url[-1] == "/"):
                    url = url[0: len(url) - 1]
                    if url == "":
                        break

            return url
        except:
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return url
    def opencode(self, file):
        """
        Open the Operation Code from designated Files
        """
        with open(file) as fin:
            data = json.loads(fin.read())
        return data

    def excel_to_list(self, file ):
        try:
            opened_wb = openpyxl.load_workbook(file)
            sheet = opened_wb.get_sheet_by_name((opened_wb.sheetnames[0]))
            n_rows = sheet.max_row
            n_columns = sheet.max_column
            start = 2
            url_list = []
            for i in range(start,n_rows+1):
                chunks = []
                for j in range(1, n_columns+1):
                    _value = sheet.cell(row=i, column=j).value
                    if isinstance(_value, str):
                        chunks.append(_value.strip())
                    else:
                        chunks.append(_value)
                url_list.append(chunks)
            self.logger.debug(url_list)
            return url_list

        except openpyxl.utils.exceptions.InvalidFileException:
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return "Invalid File Type."

        except PermissionError:
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return "Please close the file."

        except:
            self.logger.error("Below Exception Occurred!!", exc_info=True)

    def validate_redirect(self, file):
        try:
            opened_wb = openpyxl.load_workbook(file)
            sheet = opened_wb.get_sheet_by_name((opened_wb.sheetnames[0]))
            n_rows = sheet.max_row
            start = 2

            for i in range(start,n_rows+1):
                src = sheet.cell(row=i, column=1).value
                # target = src = sheet.cell(row=i, column=2).value
                target = sheet.cell(row=i, column=2).value
                status = self.check_redirect(src, target)
                # url_list.append(chunks)
                sheet.cell(row=i, column=3, value=str(status))

            opened_wb.save(file)

        except:
            self.logger.error("Below Exception Occurred!!", exc_info=True)

    def write_to_excel(self, file, list_status, is_redirect_operation):
        try:
            open_wb = openpyxl.load_workbook(file)
            sheet = open_wb.get_sheet_by_name((open_wb.sheetnames[0]))
            target_column = 3 if is_redirect_operation else 2
            sheet.cell(row=1, column=target_column, value="Status")
            for i, each in enumerate(list_status,0):
                sheet.cell(row=i+2, column=target_column, value=str(each))
            open_wb.save(file)
            return "Successfully Saved the Data."
        
        except PermissionError:
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return "Please Close the file or you don't\nhave persmission for the operation."

        except:
            open_wb.save(file)
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return "Exception Occured while performing the operation.\nPlease check Logs for more information"

    def check_redirect(self, src, target):
        try:
            output = requests.get(src, timeout=self.timeout, headers=self.headers)
            sleep(self.sleeptime/2)
            redirected_url = output.url
            cleaned_target = self.remove_trailing_slashes(target)
            cleaned_redirected_url = self.remove_trailing_slashes(redirected_url)
            status = cleaned_target == cleaned_redirected_url
            return status
        except:
            self.logger.error("Below Exception Occurred!!", exc_info=True)
            return False
        
    def get_url_status(self, url):
        """
        Get the URLs response status.
        """
        try:
            resp = requests.get(url, headers=self.headers)
            final_output = self.validate_history_and_get_status(resp, url)

            return final_output
        except requests.exceptions.MissingSchema:
            self.logger.error("Below Exceptio occurred:\n", exc_info=True)
            return {
                "status": 999,
                "message": "Missing Schema(http/https)"
            }
        except requests.exceptions.ConnectionError:
            self.logger.error("Below Exceptio occurred:\n", exc_info=True)
            return {
                "status": 999,
                "message": f"Failed to establish a new connection: {url}"
            }
        except:
            self.logger.error("Below Exceptio occurred:\n", exc_info=True)
            return {
                "status": 999,
                "message": "Unknown Exception."
            }
        
    def validate_history_and_get_status(self, resp, url):
        """
        Read the history and determine the exact URL status
        """
        try:
            url_without_schema = self.remove_trailing_slashes(url.replace("http://", "https://"))
            url_from_resp = resp.url
            url_from_resp_without_schema = self.remove_trailing_slashes(url_from_resp.replace("http://", "https://"))
            if url_from_resp_without_schema != url_without_schema:

                histories = resp.history
                
                final_status = resp.status_code
                if bool(histories):
                    for history in histories:
                        st = history.status_code
                        msg = history.reason
                        hist_url = history.url
                        hist_url_withour_schema = hist_url.replace("http://", "https://")
                        if hist_url_withour_schema != url_without_schema:
                            self.logger.debug("History URL: %s", hist_url_withour_schema)
                            break
                    final_status = st
                return {"status":final_status, "message": f"{final_status} - {msg}"}
            else:
                final_status = resp.status_code
                msg = f"{final_status} - {resp.reason}"
                return {"status":final_status, "message": msg}
        except:
            self.logger.error("Below Exceptio occurred:\n", exc_info=True)
            return 999