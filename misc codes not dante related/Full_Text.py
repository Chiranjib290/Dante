from openpyxl import Workbook
import requests
from requests.auth import HTTPBasicAuth
from urllib.parse import urlencode, quote_plus

# Configuration (taken from the request details you supplied)
AEM_BASE = "https://dpe.pwc.com"
AEM_BASE = "http://10.248.53.101:4502"
QUERY_ENDPOINT = "/crx/de/query.jsp"
USERNAME = "chiranjib.bhattacharyya@in.pwc.com"
PASSWORD = "Change@123456"
DEFAULT_HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "User-Agent": "python-requests/2.x"
}

def _build_query_params(sql_stmt, dc=None, charset="utf-8", type_="sql", showResults="true"):
    params = {
        "_dc": dc if dc is not None else "0",
        "_charset_": charset,
        "type": type_,
        "stmt": sql_stmt,
        "showResults": showResults
    }
    # return encoded query string (requests will encode when using params argument)
    return params

def _clean_path(path):
    """
    Split the path on the first occurrence of '/jcr:content' and return the left part.
    If '/jcr:content' is not present return the original path unchanged.
    """
    print(path)
    marker = "/jcr:content/"
    if marker in path:
        return path.split(marker, 1)[0]
    return path

def query_aem(sql_stmt, dc_value=None, timeout=30):
    """
    Execute the SQL query against the AEM CRX Query servlet and return a list of cleaned paths.
    Each returned path is the portion before '/jcr:content' when present.
    """
    url = AEM_BASE.rstrip("/") + QUERY_ENDPOINT
    params = _build_query_params(sql_stmt, dc=dc_value)
    try:
        resp = requests.get(
            url,
            params=params,
            headers=DEFAULT_HEADERS,
            auth=HTTPBasicAuth(USERNAME, PASSWORD),
            timeout=timeout,
            verify=True
        )
    except requests.RequestException as e:
        raise RuntimeError(f"HTTP request failed: {e}")
    if resp.status_code != 200:
        raise RuntimeError(f"AEM returned status {resp.status_code}: {resp.text[:500]}")
    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError("Response is not valid JSON")

    # Expecting structure like: { results: [ { path: "...", type: "cq:Page" }, ... ], total: N, success: true, ... }
    results = data.get("results") or []
    cleaned = []
    for item in results:
        raw_path = item.get("path")
        if not raw_path:
            continue
        cleaned_path = _clean_path(raw_path)
        cleaned.append(cleaned_path)
    return cleaned

def first_call(territory, text):
    return f"select * from nt:base where jcr:path like '/content/pwc/{territory}/%' and contains(*, '{text}')"

def xfs(territory, text):
    return f"select * from nt:base where jcr:path like '/content/experience-fragments/pwc/{territory}/%' and contains(*, '{text}')"

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)

# Example usage
if __name__ == "__main__":
    territory = "us"
    text,out = '"/content/pwc/us/en/industries/energy-utilities-resources.html"', "prod1.xlsx"
    text,out = '"/content/pwc/us/en/industries/industrial-products.html"', "prod2.xlsx"
    results = []
    results.append(query_aem(first_call(territory,text)))
    for xf in query_aem(xfs(territory,text)): 
        results.append(query_aem(first_call(territory,xf)))
    ans = set()
    for r in results:
        for path in r:
            ans.add(path)
    list_to_excel(sorted(ans),out)
    print("COMPLETED SUCCESSFULLY")