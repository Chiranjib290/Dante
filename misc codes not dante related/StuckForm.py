import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import requests
from openpyxl import Workbook
from urllib.parse import quote

#from email_validator import validate_email, EmailNotValidError
# def bool_is_email(identifier) -> bool:
#     url = f"https://dpe.pwc.com{identifier}.json"
#     try:
#         resp = requests.get(url, auth=AUTH, timeout=5)
#         resp.raise_for_status()
#         data = resp.json()
#     except requests.RequestException as e:
#         print(f"Request failed for {url}: {e}")
#         return False
#     except ValueError:
#         print(f"Invalid JSON at {url}")
#         return False
#     email_address = data.get("email")
#     if not email_address:
#         print(f"No email found in response for {identifier}")
#         return False
#     try:
#         validate_email(email_address, check_deliverability=True)
#         return True
#     except EmailNotValidError as e:
#         print(f"Invalid email '{email_address}' for {identifier}: {e}")
#         return False

def is_payload_valid(payload):   
    try:
        _content_root = "/content/pwc"
        _content_root_experience_fragment ="/content/experience-fragments/pwc"
        _content_root_dam = "/content/dam/pwc"
        _form_path =  "/content/usergenerated/content/pwc"
        _form_archive_path =  "/content/usergenerated/archive/content/pwc"
        if payload.lower().startswith(_content_root) or payload.lower().startswith(_content_root_experience_fragment) or payload.lower().startswith(_content_root_dam) or payload.lower().startswith(_form_path) or payload.lower().startswith(_form_archive_path) or payload.lower().startswith("/content/"):
            pageinfo="https://dpe.pwc.com/libs/wcm/core/content/pageinfo.json?path="+payload
            requestedpage=requests.get(pageinfo,auth=AUTH)
            return requestedpage.status_code  
        else:
            return 404
    except Exception as e:
        print("Below Exception occurred\n", e)
        return 999 
    

def spam_check_v2(process):
    for p in process:
        post_data = {'model': '/var/workflow/models/pwc-form-submission-spam-check-v2', 
                    '_charset_': 'utf-8', 
                    'payload': str(p), 
                    'payloadType': 'JCR_PATH'}
        post_url = 'https://dpe.pwc.com/var/workflow/instances'

        if is_payload_valid(p):
            post_resp_data = requests.post(post_url, data=post_data, auth=AUTH)
            if post_resp_data.status_code >= 200 and post_resp_data.status_code < 207:
                print("Payload Ran Successfully")
            else:
                print(f"Something went wrong : {post_resp_data.status_code}")


def create_sheet_1(excel_filename="paths.xlsx", wb=Workbook()):
    territories = ['US',
    'UK',
    'GX',
    'CH',
    'CA',
    'CN',
    'IN',
    'PK',
    'AU',
    'AS',
    'AR',
    'BR',
    'DN',
    'ES',
    'FR',
    'FI',
    'NO',
    'ZA']

    total_count=0
    process = []
    for t in territories:
        URL = f"https://dpe.pwc.com/bin/querybuilder.json?1_property=formtoprocess%20&1_property.value=true%20&2_property=status%20&2_property.operation=exists%20&2_property.value=false%20&3_property=jcr%3apath%20&3_property.operation=like&3_property.value=%25%2f{today_year}%2f{today_month}%2f%25%20&p.hits=selective%20&p.limit=-1%20&p.properties=jcr%3apath%20&path=%2fcontent%2fusergenerated%2fcontent%2fpwc%2f{t}%20"
        #print(URL)
        resp = requests.get(URL, auth=AUTH)
        resp.raise_for_status()
        hits = resp.json().get("hits", [])
        for i in hits: process.append(i["jcr:path"])
        total_count += len(hits)
    
    # create workbook & sheet
    ws = wb.create_sheet(title="Sheet1")
    
    # 1st row, 1st col
    ws.cell(row=1, column=1, value="1. Submission without status :")
    
    # 2nd row, 1st col with the query (inject date)
    query = f"SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status] IS NULL AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'"
    ws.cell(row=2, column=1, value=query)
    
    #4th row
    ws.cell(row=4, column=1, value="Note - Ran this query individually for the territories.")
    
    # 6th row
    ws.cell(row=6, column=1, value="Result :  -")
    ws.cell(row=6, column=2, value=f"{total_count}")
    
    # Starting at 8th row, 3rd column, write each path and sequence
    start_row = 8
    seq_col = 1
    data_col = 3
    
    for idx, d in enumerate(hits, start=1):
        row = start_row + idx - 1
        ws.cell(row=row, column=seq_col, value=idx)
        ws.cell(row=row, column=data_col, value=d["jcr:path"])
    
    # Save
    wb.save(excel_filename)
    print(f"Sheet1 written to {excel_filename}")

    print("\n\nSheet1 processing starting.....")
    spam_check_v2(process)
    print("Sheet1 processed successfully") 
    return f'''
        1. Submission without status :

        SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status] IS NULL AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'

        Note - Ran this query individually for the territories.
        Result : - {total_count} Forms\n\n
        '''

def create_sheet_2(excel_filename="paths.xlsx", wb=Workbook()):
    URL = f"https://dpe.pwc.com/bin/querybuilder.json?1_property=formtoprocess&1_property.value=true&2_property=status&2_property.value=Ready%20for%20processing&3_property=jcr%3apath&3_property.operation=like&3_property.value=%25%2f{today_year}%2f{today_month}%2f%25&group.4_property=jcr%3apath&group.4_property.operation=like&group.4_property.value=%2fcontent%2fusergenerated%2fcontent%2fpwc%2fgx%2fen%2fservices%2fpeople-organisation%2fpublications%2fworkforce-of-the-future%2fquiz%2f%25&group.5_property=jcr%3apath&group.5_property.operation=like&group.5_property.value=%2fcontent%2fusergenerated%2fcontent%2fpwc%2fgx%2fen%2fservices%2fworkforce%2fpublications%2fworkforce-of-the-future%2fquiz%2f%25&group.p.not=true&group.p.or=true&p.hits=selective&p.limit=-1&p.properties=jcr%3apath&path=%2fcontent%2fusergenerated%2fcontent%2fpwc"
    #print(URL)
    resp = requests.get(URL, auth=AUTH)
    resp.raise_for_status()
    hits = resp.json().get("hits", [])
    
    # compute counts
    total_count = len(hits)
    today_count = sum(1 for d in hits if f"/{today_year}/{today_month}/{today_date}/" in d["jcr:path"])
    pre_process = [d["jcr:path"] for d in hits if f"/{today_year}/{today_month}/{today_date}/" not in d["jcr:path"]]

    # create workbook & sheet
    ws = wb.create_sheet(title="Sheet2")
    
    # 1st row, 1st col
    ws.cell(row=1, column=1, value="2. Submissions stuck in Ready for Processing :")
    
    # 2nd row, 1st col with the query (inject date)
    query = f"SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true'  AND comp.[status]='Ready for processing' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%' AND comp.[jcr:path] NOT LIKE '/content/usergenerated/content/pwc/gx/en/services/people-organisation/publications/workforce-of-the-future/quiz/%' AND comp.[jcr:path] NOT LIKE '/content/usergenerated/content/pwc/gx/en/services/workforce/publications/workforce-of-the-future/quiz/%'"
    ws.cell(row=2, column=1, value=query)
    
    # 4th row
    ws.cell(row=4, column=1, value="Result :  -")
    ws.cell(row=4, column=2, value=f"{total_count} - {today_count}")
    ws.cell(row=4, column=3, value="forms submission paths from today")
    
    # Starting at 6th row, 3rd column, write each path and sequence
    start_row = 6
    seq_col = 1
    data_col = 3
    
    for idx, d in enumerate(hits, start=1):
        row = start_row + idx - 1
        ws.cell(row=row, column=seq_col, value=idx)
        ws.cell(row=row, column=data_col, value=d["jcr:path"])
    
    # Save
    wb.save(excel_filename)
    print(f"Sheet2 written to {excel_filename}")

    print("\n\nSheet2 processing starting.....")
    process = []
    for i in pre_process:
        router_present = False
        router_url = f"https://dpe.pwc.com/bin/querybuilder.json?1_group.property=status&1_group.property.value=ACTIVE&2_property=contentPath&2_property.value={quote(i, safe='')}&path=%2Fvar%2Fworkflow%2Finstances%2Fserver1"
        router_resp = requests.get(router_url, auth=AUTH)
        if(router_resp.status_code == 200):
            len_of_fetched_router = len(router_resp.json()["hits"])
            if len_of_fetched_router > 0:
                router_present = True
        if router_present:
            node_available = requests.get('https://dpe.pwc.com' + i + ".json", auth=AUTH)
            if(node_available.status_code == 200):
                post_data = {'runModes': ['crx3tar', 'publish', 'nosamplecontent', 'crx3', 's7connect', 'prod']}
                post_resp_runmode = requests.post('https://dpe.pwc.com' + i, data=post_data, auth=AUTH)
                if post_resp_runmode.status_code == 200:
                    post_data = {'model': '/var/workflow/models/pwc-form-email', 
                    '_charset_': 'utf-8', 
                    'payload': str(i), 
                    'payloadType': 'JCR_PATH'}
                    post_url = 'https://dpe.pwc.com/var/workflow/instances'
                    if is_payload_valid(i):
                        post_resp_data = requests.post(post_url, data=post_data, auth=AUTH)
                        if post_resp_data.status_code >= 200 and post_resp_data.status_code < 207:
                            print("Payload Ran Successfully")
                        else:
                            print(f"Something went wrong : {post_resp_data.status_code}")
        else:
            process.append(i)
    spam_check_v2(process)
    print("Sheet2 processed successfully") 

    return f'''
        2. Submissions stuck in Ready for Processing :

        SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true'  AND comp.[status]='Ready for processing' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%' AND comp.[jcr:path] NOT LIKE '/content/usergenerated/content/pwc/gx/en/services/people-organisation/publications/workforce-of-the-future/quiz/%' AND comp.[jcr:path] NOT LIKE '/content/usergenerated/content/pwc/gx/en/services/workforce/publications/workforce-of-the-future/quiz/%'

        Result : {total_count} - {today_count} forms submission paths from today\n\n
        '''

def create_sheet_3(excel_filename="paths.xlsx", wb=Workbook()):
    URL = f"https://dpe.pwc.com/bin/querybuilder.json?1_property=formtoprocess&1_property.value=true&2_property=status&2_property.value=Processed&3_property=jcr%3apath&3_property.operation=like&3_property.value=%25%2f{today_year}%2f{today_month}%2f%25&p.hits=selective&p.limit=-1&p.properties=jcr%3apath&path=%2fcontent%2fusergenerated%2fcontent%2fpwc"
    #print(URL)
    resp = requests.get(URL, auth=AUTH)
    resp.raise_for_status()
    hits = resp.json().get("hits", [])
    
    # compute counts
    total_count = len(hits)
    today_count = sum(1 for d in hits if f"/{today_year}/{today_month}/{today_date}/" in d["jcr:path"])
    pre_process = [d["jcr:path"] for d in hits if f"/{today_year}/{today_month}/{today_date}/" not in d["jcr:path"]]

    # create workbook & sheet
    ws = wb.create_sheet(title="Sheet3")
    
    # 1st row, 1st col
    ws.cell(row=1, column=1, value="3. Submission stuck in Processed :")
    
    # 2nd row, 1st col with the query (inject date)
    query = f"SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Processed' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%' "
    ws.cell(row=2, column=1, value=query)
    
    # 4th row
    ws.cell(row=4, column=1, value="Result :  -")
    ws.cell(row=4, column=2, value=f"{total_count} - {today_count}")
    ws.cell(row=4, column=3, value="forms submission paths from today")
    
    # Starting at 6th row, 3rd column, write each path and sequence
    start_row = 6
    seq_col = 1
    data_col = 3
    
    for idx, d in enumerate(hits, start=1):
        row = start_row + idx - 1
        ws.cell(row=row, column=seq_col, value=idx)
        ws.cell(row=row, column=data_col, value=d["jcr:path"])
    
    # Save
    wb.save(excel_filename)
    print(f"Sheet3 written to {excel_filename}")

    print("\n\nSheet3 processing starting.....")
    process = []
    for i in pre_process:
        post_url = f'https://dpe.pwc.com{i}'
        post_data = {'status': 'Submitted', 'email@TypeHint': 'String'}
        output = requests.post(post_url, data=post_data, auth=AUTH)
        if output.status_code == 200:
            process.append(i)
    spam_check_v2(process)
    print("Sheet3 processed successfully") 

    return f'''
        3. Submission stuck in Processed :

        SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Processed' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%' 

        Result :  {total_count} - {today_count} forms submission paths from today.\n\n
        '''
    
def create_sheet_4(excel_filename="paths.xlsx", wb=Workbook()):
    URL = f"https://dpe.pwc.com/bin/querybuilder.json?1_property=formtoprocess&1_property.value=true&2_property=status&2_property.value=Submitted&3_property=jcr%3apath&3_property.operation=like&3_property.value=%25%2f{today_year}%2f{today_month}%2f%25&p.hits=selective&p.limit=-1&p.properties=jcr%3apath&path=%2fcontent%2fusergenerated%2fcontent%2fpwc"
    #print(URL)
    resp = requests.get(URL, auth=AUTH)
    resp.raise_for_status()
    hits = resp.json().get("hits", [])
    
    # compute counts
    total_count = len(hits)
    today_count = sum(1 for d in hits if f"/{today_year}/{today_month}/{today_date}/" in d["jcr:path"])
    process = [d["jcr:path"] for d in hits if f"/{today_year}/{today_month}/{today_date}/" not in d["jcr:path"]]
    # create workbook & sheet
    ws = wb.create_sheet(title="Sheet4")
    
    # 1st row, 1st col
    ws.cell(row=1, column=1, value="4. Submission stuck in Submitted :")
    
    # 2nd row, 1st col with the query (inject date)
    query = f"SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Submitted' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'"
    ws.cell(row=2, column=1, value=query)
    
    # 4th row
    ws.cell(row=4, column=1, value="Result :  -")
    ws.cell(row=4, column=2, value=f"{total_count} - {today_count}")
    ws.cell(row=4, column=3, value="forms submission paths from today")
    
    # Starting at 6th row, 3rd column, write each path and sequence
    start_row = 6
    seq_col = 1
    data_col = 3
    
    for idx, d in enumerate(hits, start=1):
        row = start_row + idx - 1
        ws.cell(row=row, column=seq_col, value=idx)
        ws.cell(row=row, column=data_col, value=d["jcr:path"])
    
    # Save
    wb.save(excel_filename)
    print(f"Sheet4 written to {excel_filename}")
    print("\n\nSheet4 processing starting.....")
    spam_check_v2(process)
    print("Sheet4 processed successfully") 
    return f'''
        4. Submission stuck in Submitted:

        SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Submitted' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'

        Result : {total_count}-{today_count} forms submission paths from today.\n\n

        '''
 
def create_sheet_5(excel_filename="paths.xlsx", wb=Workbook()):
    URL = f"https://dpe.pwc.com/bin/querybuilder.json?1_property=formtoprocess&1_property.value=true&2_property=status&2_property.value=Completed-Banned-Words&3_property=jcr%3apath&3_property.operation=like&3_property.value=%25%2f{today_year}%2f{today_month}%2f%25&p.hits=selective&p.limit=-1&p.properties=jcr%3apath&path=%2fcontent%2fusergenerated%2fcontent%2fpwc"
    #print(URL)
    resp = requests.get(URL, auth=AUTH)
    resp.raise_for_status()
    hits = resp.json().get("hits", [])
    
    # compute counts
    total_count = len(hits)
    
    # create workbook & sheet
    ws = wb.create_sheet(title="Sheet5")
    
    # 1st row, 1st col
    ws.cell(row=1, column=1, value="5. Submission stuck with Completed-Banned-Words status :")
    
    # 2nd row, 1st col with the query (inject date)
    query = f"SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Completed-Banned-Words' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'"
    ws.cell(row=2, column=1, value=query)
    
    # 4th row
    ws.cell(row=4, column=1, value="Result :  -")
    ws.cell(row=4, column=2, value=f"{total_count}")
    ws.cell(row=4, column=3, value="Forms")
    
    # Starting at 6th row, 3rd column, write each path and sequence
    start_row = 6
    seq_col = 1
    data_col = 3
    
    for idx, d in enumerate(hits, start=1):
        row = start_row + idx - 1
        ws.cell(row=row, column=seq_col, value=idx)
        ws.cell(row=row, column=data_col, value=d["jcr:path"])
    
    # Save
    wb.save(excel_filename)
    print(f"Sheet5 written to {excel_filename}")
    
    print("\n\nSheet5 processing starting.....")
    process = []
    for i in hits:
        i = i["jcr:path"]
        #if bool_is_email(i):

        get_url= f'https://dpe.pwc.com{i}.json'
        resp_each = requests.get(get_url, auth=AUTH)
        if resp_each.status_code == 200:
            email = resp_each.json().get('email','')
            fixed_email=email.strip()
            post_url = f'https://dpe.pwc.com{i}'
            post_data = {'email': fixed_email, 'email@TypeHint': 'String'}
            output = requests.post(post_url, data=post_data, auth=AUTH)
            if output.status_code == 200:
                process.append(i)
    spam_check_v2(process)
    print("Sheet5 processed successfully") 
    return f'''
        5. Submission stuck with completed Banned words status:

        SELECT [jcr:path] FROM [nt:unstructured] AS comp WHERE ISDESCENDANTNODE(comp, '/content/usergenerated/content/pwc') AND [formtoprocess]='true' AND comp.[status]='Completed-Banned-Words' AND comp.[jcr:path] LIKE '%/{today_year}/{today_month}/%'

        Result :  {total_count} forms.\n\n
        '''


today = datetime.date.today()
today_date = today.day
today_month = today.month
today_year = today.year

# Format date for filename and email subject
excel_filename = f"Stuck Form Report {today_date:02d}-{today_month:02d}-{today_year}.xlsx"
email_subject = f"Stuck Form Processing Report - {today_month:02d}/{today_date:02d}/{today_year}"
email_body = "Hello Everyone,\n\n"

AUTH = ("chiranjib.bhattacharyya@in.pwc.com", "Change@123456")
excel_name = f"Stuck Form Report {today_date:02d}-{today_month:02d}-{today_year}.xlsx"
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet created by Workbook()

email_body+=create_sheet_1(excel_filename=excel_name, wb=wb)
email_body+=create_sheet_2(excel_filename=excel_name, wb=wb)
email_body+=create_sheet_3(excel_filename=excel_name, wb=wb)
email_body+=create_sheet_4(excel_filename=excel_name, wb=wb)
email_body+=create_sheet_5(excel_filename=excel_name, wb=wb)

wb.save(excel_name)
print(f"All sheets written to {excel_name}")

email_body+='''
(Kindly go through the attached report for further details).

Thanks and Regards,
Chiranjib Bhattacharyya
'''
print(email_body)
# Email setup
sender_email = "chiranjib290@gmail.com"
password = "gmjp gkyo ebys qttn"
recipients = [
    "chiranjib.bhattacharyya@pwc.com"
]

# Construct the email
message = MIMEMultipart()
message["From"] = sender_email
message["To"] = ", ".join(recipients)
message["Subject"] = email_subject
message.attach(MIMEText(email_body, "plain"))

# Attach the Excel file
with open(excel_filename, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={excel_filename}")
    message.attach(part)

# Send the email
try:
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender_email, password)
    server.sendmail(sender_email, recipients, message.as_string())
    server.quit()
    print("Email sent successfully.")
except Exception as e:
    print("Failed to send email:", e)
