import openpyxl
from openpyxl import Workbook

# Define the common prefix that we want to remove from each path.
COMMON_PREFIX = "/content/cq:tags/pwc-pl/"

def insert_path(root, path_parts, title):
    """
    Inserts a tokenized path into the tree along with its title.
    
    Parameters:
    - root: The current level of the tree (a dict).
    - path_parts: List of path tokens (folder names).
    - title: The title for the final token.
    """
    if not path_parts:
        return

    # Token for the current level
    key = path_parts[0]

    # If this token is not already present in the tree, add it with a placeholder for its children.
    if key not in root:
        # Use the provided title only if this is the last token; otherwise, use the key as the title.
        node_title = title if len(path_parts) == 1 else key
        root[key] = {"title": node_title, "children": {}}

    # Recursively insert any remaining tokens in the path.
    insert_path(root[key]["children"], path_parts[1:], title)

def build_tree_from_excel(filename):
    """
    Reads the Excel file where:
    - Column A contains the full path.
    - Column B contains the title.
    
    It returns a nested dictionary representing the folder hierarchy.
    """
    tree = {}
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Iterate through each row. If there is a header, adjust the iteration accordingly.
    for row in ws.iter_rows(values_only=True):
        # Skip empty rows
        if not row or row[0] is None or row[1] is None:
            continue

        full_path, title = row[0].strip(), row[1].strip()

        # Remove the common prefix
        if full_path.startswith(COMMON_PREFIX):
            relative_path = full_path[len(COMMON_PREFIX):]
        else:
            relative_path = full_path

        # Tokenize the path; filter out empty strings (if any)
        path_tokens = [token for token in relative_path.split("/") if token]
        insert_path(tree, path_tokens, title)

    return tree

def traverse_tree(tree, level=0, current_row=None):
    """
    Performs a depth-first traversal of the tree.
    Each node's title is placed in a cell corresponding to its depth (level) in the folder hierarchy.
    
    Returns:
    A list of rows where each row is a list representing the hierarchy, 
    e.g., [level1_title, level2_title, level3_title, ...].
    """
    if current_row is None:
        current_row = []
    
    rows = []
    # Process nodes in sorted order for consistency.
    for key in sorted(tree.keys()):
        node = tree[key]
        row = current_row[:]
        if len(row) <= level:
            row.extend([None] * (level - len(row) + 1))
        row[level] = node["title"]
        rows.append(row)
        if node["children"]:
            child_rows = traverse_tree(node["children"], level=level + 1, current_row=row)
            rows.extend(child_rows)
    return rows

def write_tree_to_excel(rows, output_filename):
    """
    Writes the hierarchy rows to an Excel file.
    Each column represents a level in the folder structure.
    
    Parameters:
    - rows: List of rows representing the folder hierarchy.
    - output_filename: Name of the Excel file to save.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Tree"

    for i, row in enumerate(rows, start=1):
        for j, cell_value in enumerate(row, start=1):
            if cell_value is not None:
                ws.cell(row=i, column=j, value=cell_value)

    wb.save(output_filename)
    print(f"Excel file saved as: {output_filename}")

if __name__ == "__main__":
    # Build the tree using the 'pl_tags.xlsx' file.
    input_filename = "pl_tags.xlsx"
    tree = build_tree_from_excel(input_filename)

    # Traverse the tree to generate rows corresponding to the folder hierarchy.
    rows = traverse_tree(tree)

    # Write the hierarchy to an Excel file.
    output_filename = "tags_hierarchy.xlsx"
    write_tree_to_excel(rows, output_filename)
