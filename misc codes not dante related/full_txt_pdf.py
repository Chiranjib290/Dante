#!/usr/bin/env python3
# Requires: pip install requests PyPDF2 pycryptodome
from openpyxl import Workbook
import requests
from requests.auth import HTTPBasicAuth
from io import BytesIO
from PyPDF2 import PdfReader
import time

# Configuration
base_host = "https://dpe.pwc.com"
territory="de"
search_substring = "cyber"
out="prod_output_pdfs.xlsx"

query_endpoint = f"/bin/querybuilder.json?group.1_property=jcr%3apath&group.1_property.operation=like&group.1_property.value=%25.pdf&p.hits=selective&p.limit=-1&p.properties=jcr%3apath&path=%2fcontent%2fdam%2fpwc%2f{territory}%2f&type=dam%3aAsset"
query_url = base_host + query_endpoint

username = "chiranjib.bhattacharyya@in.pwc.com"
password = "Change@123456"

per_request_delay = 0.2

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)

def fetch_json(url, auth, timeout=30):
    r = requests.get(url, auth=auth, timeout=timeout)
    r.raise_for_status()
    return r.json()


def list_jcr_paths(query_json):
    paths = []
    if isinstance(query_json, dict) and "hits" in query_json:
        for hit in query_json["hits"]:
            if isinstance(hit, dict) and "jcr:path" in hit:
                paths.append(hit["jcr:path"])
    return paths


def fetch_pdf_stream(pdf_url, auth, timeout=60):
    with requests.get(pdf_url, auth=auth, stream=True, timeout=timeout) as r:
        r.raise_for_status()
        buf = BytesIO()
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:
                buf.write(chunk)
        buf.seek(0)
        return buf


def extract_text_from_pdf_stream(stream):
    """
    Attempt to read text from a PDF bytes stream using PyPDF2.
    Handles encrypted PDFs by attempting decryption with an empty password.
    Returns the extracted text (possibly empty) and a status string.
    """
    try:
        reader = PdfReader(stream)
    except Exception as e:
        return "", f"reader_init_error: {e}"

    # handle encrypted PDFs
    if getattr(reader, "is_encrypted", False):
        try:
            # Try decrypt with empty password
            decrypt_result = reader.decrypt("")  # returns 0 for failure, >0 for success in many PyPDF2 variants
            if not decrypt_result:
                # If decrypt returns falsy, try common fallback - None/0 means failed.
                return "", "encrypted_unrecoverable"
        except Exception as e:
            return "", f"decrypt_error: {e}"

    texts = []
    try:
        for page in reader.pages:
            try:
                txt = page.extract_text()
            except Exception:
                txt = None
            if txt:
                texts.append(txt)
    except Exception as e:
        return "", f"page_iter_error: {e}"

    return "\n".join(texts), "ok"


def pdf_contains_substring(pdf_url, auth, substring, case_insensitive=True):
    stream = fetch_pdf_stream(pdf_url, auth)
    text, status = extract_text_from_pdf_stream(stream)
    if status != "ok":
        # Return status so caller knows why it wasn't searchable
        return False, status
    if case_insensitive:
        return (substring.lower() in text.lower()), "ok"
    return (substring in text), "ok"


def main():
    auth = HTTPBasicAuth(username, password)
    print("Fetching QueryBuilder JSON from:", query_url)
    try:
        qjson = fetch_json(query_url, auth)
    except Exception as e:
        print("Failed to fetch query JSON:", e)
        return

    paths = list_jcr_paths(qjson)
    if not paths:
        print("No jcr:paths found in query response.")
        return
    paths_check = ['/content/dam/pwc/de/de/im-fokus/cyber-security-privacy/cyber-risk-management.pdf','/content/dam/pwc/de/de/strategie-organisation-prozesse-systeme/pwc-germany-cyber-escape-room-customer-product-presentation-english.pdf']
    check = 0
    print(f"Found {len(paths)} paths. Searching for substring: '{search_substring}'")

    matching_paths = []
    for idx, p in enumerate(paths, start=1):
        pdf_url = base_host.rstrip("/") + p
        try:
            found, status = pdf_contains_substring(pdf_url, auth, search_substring, case_insensitive=True)
            if found:
                matching_paths.append(p)
                print(f"[{idx}/{len(paths)}] MATCH: {p}")
                if p in paths_check:
                    check +=1
                    print('\n\n')
            else:
                if status == "ok":
                    print(f"[{idx}/{len(paths)}] no match: {p}")
                else:
                    print(f"[{idx}/{len(paths)}] ERROR ({status}) for {p}")
        except Exception as e:
            print(f"[{idx}/{len(paths)}] EXCEPTION fetching/extracting {p}: {e}")
        time.sleep(per_request_delay)

    print("\nMatching PDF jcr:paths:")
    for mp in matching_paths:
        print(mp)
    print(f"\nTotal matches: {len(matching_paths)}")
    list_to_excel(sorted(matching_paths),out)
    if check == len(paths_check):
        print('Correct Result')
    else:
        print('Incorrect Result')

if __name__ == "__main__":
    main()
