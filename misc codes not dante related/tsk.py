import requests
from openpyxl import load_workbook
import concurrent.futures
 
 
 
input_file_name="bs_sub 2.xlsx"
wb = load_workbook(input_file_name)
ws = wb.active
 
uname="admin"
passwd="BSpr0d@A3m2019"
 
row=1
col=0
 
 
def fetch_and_update(i, value):
    print(i)
    print(value)
    response = requests.get(f"http://10.195.129.68:4503{value}.json", auth=(uname, passwd))
    if response.status_code == 200:
        response_2 = requests.get(f"http://10.195.129.69:4503{value}.json", auth=(uname, passwd))
        if response_2.status_code == 200:
            output = "User Present in Pub1"
        else:
            output = "User Not Present in Pub1"        
           
    else:
        output = f"Error: {response.status_code}"
       
 
   
    # Update the worksheet cell
    ws.cell(row=i, column=4, value=output)
   
 
 
# Use ThreadPoolExecutor to fetch data concurrently
with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
    futures = []
    for i, cell in enumerate(ws['A'][1:], start=2):
        value = cell.value
        futures.append(executor.submit(fetch_and_update, i, value))
 
    # Optionally wait for all futures to complete
    concurrent.futures.wait(futures)
 
 
 
wb.save(input_file_name)