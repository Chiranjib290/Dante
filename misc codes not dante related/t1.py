from openpyxl import Workbook
import requests
from requests.auth import HTTPBasicAuth

# Configuration (as you requested)
AEM_BASE = "http://10.248.54.68:4502"
QUERY_ENDPOINT = "/crx/de/query.jsp"
USERNAME = "chiranjib.bhattacharyya@in.pwc.com"
PASSWORD = "Change@123456"
DEFAULT_HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "User-Agent": "python-requests/2.x"
}

def _build_query_params(sql_stmt, dc=None, charset="utf-8", type_="sql", showResults="true"):
    return {
        "_dc": dc if dc is not None else "0",
        "_charset_": charset,
        "type": type_,
        "stmt": sql_stmt,
        "showResults": showResults
    }

def _clean_path(path):
    """
    Split the path on the first occurrence of '/jcr:content' (no trailing slash) and return the left part.
    If '/jcr:content' is not present return the original path unchanged.
    """
    marker = "/jcr:content"
    if marker in path:
        return path.split(marker, 1)[0]
    return path

def _escape_contains_value(value):
    # Basic escaping for single quotes inside the contains() argument
    return value.replace("'", "''")

def query_aem(sql_stmt, dc_value=None, timeout=30, verify=True):
    url = AEM_BASE.rstrip("/") + QUERY_ENDPOINT
    params = _build_query_params(sql_stmt, dc=dc_value)
    try:
        resp = requests.get(
            url,
            params=params,
            headers=DEFAULT_HEADERS,
            auth=HTTPBasicAuth(USERNAME, PASSWORD),
            timeout=timeout,
            verify=verify
        )
    except requests.RequestException as e:
        raise RuntimeError(f"HTTP request failed: {e}")
    if resp.status_code != 200:
        raise RuntimeError(f"AEM returned status {resp.status_code}: {resp.text[:500]}")
    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError("Response is not valid JSON")
    results = data.get("results") or []
    cleaned = []
    for item in results:
        raw_path = item.get("path")
        if not raw_path:
            continue
        cleaned_path = _clean_path(raw_path)
        cleaned.append(cleaned_path)
    return cleaned

def first_call(territory, text):
    safe = _escape_contains_value(text)
    return f"select * from nt:base where jcr:path like '/content/pwc/{territory}/%' and contains(*, '{safe}')"

def xfs(territory, text):
    safe = _escape_contains_value(text)
    return f"select * from nt:base where jcr:path like '/content/experience-fragments/pwc/{territory}/%' and contains(*, '{safe}')"

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)

if __name__ == "__main__":
    territory = "de"
    search_text = "Trust in What Matters"  # plain text; it's escaped before inclusion in queries
    results = []

    # top-level pages matching the search_text
    top_pages = query_aem(first_call(territory, search_text), verify=True)
    results.extend(top_pages)

    # experience fragments that match the search_text
    xfragments = query_aem(xfs(territory, search_text), verify=True)

    # For each fragment found, search pages using the fragment's final segment as a search term
    for frag_path in xfragments:
        fragment_name = frag_path.rsplit("/", 1)[-1] if frag_path else frag_path
        pages_for_fragment = query_aem(first_call(territory, fragment_name), verify=True)
        results.extend(pages_for_fragment)

    ans = set(results)
    list_to_excel(sorted(ans), "prod_output_cb.xlsx")
    print("COMPLETED SUCCESSFULLY")
