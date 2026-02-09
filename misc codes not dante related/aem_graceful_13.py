import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import concurrent.futures


fnme = 'brandiste_tr_23_output'
input_file_name=f"C:/Users/cbhattacha015/Desktop/TST/DPE_Automation_Tool/{fnme}.xlsx"
wb = load_workbook(input_file_name)
ws = wb.active
 
uname="admin"
passwd="BSpr0d@A3m2019"
 
row=1
col=0
 
# for i, cell in enumerate(ws['A'][1:], start=2):
#     print(i)  
#     value = cell.value
#     print(value)
#     response = requests.get(f"http://10.195.129.68:4503{value}/profile.json",auth=(uname,passwd))
#     if response.status_code == 200:
#        data=response.json()
#        #print(data['lastLoginDate'])
#        if 'lastLoginDate' in data:
#         output = data['lastLoginDate'] if data['lastLoginDate'] else "Not Found"
#        else:
#         output = "Key Not Found"
#        ws.cell(row=i, column=2, value=output)
 
def fetch_and_update(i, value):
    print(i)
    print(value)
    response = requests.get(f"http://10.195.129.68:4503{value}/profile.json", auth=(uname, passwd))
    if response.status_code == 200:
        data = response.json()
        if 'lastLoginDate' in data:
            output = data['lastLoginDate'] if data['lastLoginDate'] else "Not Found"
        else:
            output = "Key Not Found"
    else:
        output = f"Error: {response.status_code}"
   
    # Update the worksheet cell
    ws.cell(row=i, column=2, value=output)
 
# Use ThreadPoolExecutor to fetch data concurrently
with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
    futures = []
    for i, cell in enumerate(ws['A'][1:], start=2):
        value = cell.value
        futures.append(executor.submit(fetch_and_update, i, value))
 
    # Optionally wait for all futures to complete
    concurrent.futures.wait(futures)
 
 
 
output_file_name = f"C:/Users/cbhattacha015/Desktop/TST/DPE_Automation_Tool/{fnme}_output.xlsx"
wb.save(output_file_name)  
