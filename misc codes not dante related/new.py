import requests
from openpyxl import load_workbook
import concurrent.futures
 
input_file_name="delete.xlsx"
 
wb = load_workbook(input_file_name)
ws = wb.active
 
BASE_URL    = 'http://10.195.129.69:4503'
BASE_URL_2    = 'http://10.195.129.68:4503'
USERNAME    = 'admin'
PASSWORD    = 'BSpr0d@A3m2019'
AUTH    = (USERNAME, PASSWORD)
 
row=1
col=0
 
 
def user_delete(i, path):
    print(path)
   
    url  = f"{BASE_URL}{path}"
    url_2 = f"{BASE_URL_2}{path}"
    data = {'deleteAuthorizable': ''}
   
    response = requests.get(f"{BASE_URL}{path}.json", auth=AUTH , timeout=20)
    if response.status_code == 200:
        response = requests.post(url, data=data, auth=AUTH, timeout=20)
        if response.status_code == 200:
          output=f"Deleted"
        else:
          output=f"Failed to delete"
    else:
       output=f"Unable to find user"
    ws.cell(row=i, column=2, value=output)  
   
    response_2 = requests.get(f"{BASE_URL_2}{path}.json", auth=AUTH , timeout=20)
    if response_2.status_code == 200:
        response = requests.post(url_2, data=data, auth=AUTH, timeout=20)
        if response.status_code == 200:
          output_2=f"Deleted"
        else:
          output_2=f"Failed to delete"
    else:
       output_2=f"Unable to find user"
    ws.cell(row=i, column=3, value=output_2)
 
with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
    futures = []
    for i, cell in enumerate(ws['A'][1:], start=2):
        value = cell.value
        futures.append(executor.submit(user_delete, i, value))
 
    concurrent.futures.wait(futures)
 
 
 
output_file_name=input_file_name
wb.save(output_file_name)  