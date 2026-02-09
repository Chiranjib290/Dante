from openpyxl import load_workbook
import requests

AUTH = ("admin","admin")
BASE_URL = "http://localhost:4502/"
if BASE_URL[-1]=='/':BASE_URL=BASE_URL[:-1]
EXCEL_FILENAME = "test.xlsx"
GROUP_NAME="test_group_add"
GROUP_PATH="/home/groups/H/HnB5IBvN25FnMiq1rjzm"
MEMBERS = requests.get(f"{BASE_URL}{GROUP_PATH}.json",auth=AUTH).json().get("rep:members")
print(MEMBERS)

def excel_to_list(excel_filename=EXCEL_FILENAME):
    wb = load_workbook(excel_filename)
    ws = wb.active
    return [cell.value for cell in ws['A'][1:]]

def get_user_id(user_path):
    url=f"{BASE_URL}{user_path}.json"
    try:
        response = requests.get(url,auth=AUTH)
        data=response.json()
        data=[data.get("rep:authorizableId"),data.get("jcr:uuid")]
        return data
    except Exception as e:
        print(f"Could Not fetch id for user {user_path}.\nRaised Exception : {e}")
        return False


def add_user_to_group(user_path):
    url = f"{BASE_URL}{GROUP_PATH}.rw.html"
    ids = get_user_id(user_path)
    au_id=ids[0]
    uuid=ids[1]
    
    if ids:
        if not MEMBERS or uuid not in MEMBERS:
            data = {
                "addMembers": au_id
            }
            try:
                response = requests.post(url, data=data, auth=AUTH, timeout=30)
            except Exception as e:
                print(f"Failed to add user '{au_id}' to group '{GROUP_NAME}': {e}")
                return False
        
            if response.status_code in (200, 201):
                print(f"Added user '{au_id}' to group '{GROUP_NAME}'.")
                return True
            else:
                print(f"Failed to add user '{au_id}' to group '{GROUP_NAME}': HTTP {response.status_code} {response.text}")
                return False
        else:
            print(f"user {au_id} or {uuid} or {user_path} already in {GROUP_NAME}")

def main():
    userpaths=excel_to_list()
    for u in userpaths:
        add_user_to_group(u)

if __name__=="__main__":
    main()