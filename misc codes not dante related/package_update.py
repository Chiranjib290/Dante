from openpyxl import load_workbook
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import time

def excel_to_list(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
    return [cell.value for cell in ws['A'][1:]]

AEM_BASE = "http://localhost:4502"
PACKAGE_NAME = "demo"
PARENT_PATH = f"/etc/packages/my_packages/{PACKAGE_NAME}.zip/jcr:content/vlt:definition/filter"
USERNAME = "admin"
PASSWORD = "admin"
EXCEL_FILE = 'test.xlsx'
USER_PATHS = excel_to_list(EXCEL_FILE)

start_index = 0

REQUEST_DELAY = 0.15

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

session = requests.Session()
session.auth = HTTPBasicAuth(USERNAME, PASSWORD)
session.verify = False


def node_exists(node_name: str) -> bool:
    check_url = AEM_BASE.rstrip("/") + PARENT_PATH + "/" + node_name + ".json"
    resp = session.get(check_url)
    return resp.status_code == 200

def create_filter_node(index: int, root: str):
    node_name = f"f{index}"
    url = AEM_BASE.rstrip("/") + PARENT_PATH  # POST to parent path

    payload = {
        "jcr:primaryType": "nt:unstructured",
        "mode": "replace",
        "root": root,
        "rules": ['exclude:/home/users/.*/.tokens'],
        "rules@TypeHint": "String[]"
    }

    resp = session.post(f"{url}/{node_name}", data=payload)
    return resp

def main():
    created = []
    skipped = []
    errors = []
    idx = start_index
    for user_root in USER_PATHS:
        node_name = f"f{idx}"
        try:
            if node_exists(node_name):
                skipped.append(node_name)
                print(f"[SKIP] {node_name} already exists; skipping.")
            else:
                resp = create_filter_node(idx, user_root)
                if resp.status_code in (200, 201, 204):
                    created.append(node_name)
                    print(f"[OK] Created {node_name} -> root={user_root} (HTTP {resp.status_code})")
                else:
                    snippet = resp.text[:300].replace("\n", " ")
                    errors.append((node_name, resp.status_code, snippet))
                    print(f"[FAIL] {node_name} HTTP {resp.status_code} - {snippet}")
        except Exception as e:
            errors.append((node_name, "exception", str(e)))
            print(f"[ERROR] {node_name} exception: {e}")

        idx += 1
        time.sleep(REQUEST_DELAY)

    print("\nSummary:")
    print(f"  Created: {len(created)} -> {created}")
    print(f"  Skipped: {len(skipped)} -> {skipped}")
    if errors:
        print(f"  Errors: {len(errors)}")
        for e in errors:
            print(f"    {e}")

if __name__ == "__main__":
    main()
