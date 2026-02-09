import requests
from openpyxl import load_workbook
 
output_file_name = "symbolic_name_output.xlsx"
input_file_name = "bundles_array.xlsx"
wb = load_workbook(input_file_name)
ws = wb.active
 
uname = "admin"
passwd = "admin"
 
response = requests.get("http://localhost:4502/system/console/bundles.json", auth=(uname, passwd))
if response.status_code == 200:
    data = response.json()
    data2 = data['data']
 
    # Start writing from the first row
    row = 2
 
    for a in data2:
        output = a['symbolicName']
        #print(output)
        ws.cell(row=row, column=1, value=output)
        row += 1
 
    # Save the workbook after writing all the data
    
    wb.save(output_file_name)
 
else:
    print(f"Failed to retrieve bundles. Status code: {response.status_code}")