from openpyxl import Workbook, load_workbook
import requests as r


AUTH = ('chiranjib.bhattacharyya@in.pwc.com','Change@123456')
AEM_BASE = "https://dpe.pwc.com"



res = []
def deletion(ccf) :
    url = f"{AEM_BASE}/bin/wcm/references.json?path={ccf}"
    resp = r.get(url, auth=AUTH)
    hits = resp.json().get('pages','[]')
    if hits!=[]:
        hits = hits[0].get('references','[]')
        for i in hits :
            print(f"  -  {i}")
            if not i.startswith("/content/pwc/sk/sk/archive-folder") and not i.startswith("/content/pwc/sk/en/archive-folder") : return False
    return True

def extraction(lang):
    url = f"{AEM_BASE}/bin/querybuilder.json?p.hits=selective&p.limit=-1&p.peoperties=jcr%3apath&path=%2fcontent%2fexperience-fragments%2fpwc%2fsk%2f{lang}%2fcontact&property=sling%3aresourceType&property.operation=like&property.value=%25experiencefragment"
    resp = r.get(url, auth=AUTH)
    hits = resp.json().get('hits','[]')
    for i in hits :
        x = i.get('jcr:path','')
        sub = x[:x.find('/jcr:content')]
        print(sub)
        chk = deletion(sub)
        print(chk)
        if chk : res.append(sub)

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)

def excel_to_list(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
    return [cell.value for cell in ws['A'][1:]]

def page_exists(content_path):
    resp = r.get(f'https://dpe.pwc.com/libs/wcm/core/content/pageinfo.json?path={content_path}', auth=AUTH)
    return resp.status_code==200

ls = excel_to_list('test.xlsx')
for url in ls:
    if not page_exists(url):
        print(url)