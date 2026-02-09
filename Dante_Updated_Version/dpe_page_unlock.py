import requests
import json
import time
import json
import logging
import xlrd
import openpyxl 


class UnlockPages:
    """Made by Debolina Dutta , Aman Pratiush, Date: 05/09/2024.
    Unlock DPE Pages in bulk
    """
    def __init__(self,ip,user,passwd):

        self.ip = ip.strip()
        self.user = user.strip()
        self.passwd = passwd.strip()

        self.configdata = self.opencode("configfiles\\config.json")
        self.loglevel = self.configdata["loglevel"]
        self.sleeptime = self.configdata["sleeptime"]
        self.timeout = self.configdata["timeout"]
        self.commonpasswordforalluser=self.configdata["commonpasswordforalluser"]

        log_level = {
            'debug': logging.DEBUG,
            'info': logging.INFO,
            'warning': logging.WARNING,
            'error': logging.ERROR,
            'critical': logging.CRITICAL
        }
        self.logger = logging.getLogger()
        self.logger.setLevel(log_level[self.loglevel])
        self.wfoperationdata = self.opencode("configfiles\\operationcode.json")
        self.logger.debug("Operation data : " + str(self.wfoperationdata))
        self.logger.debug("Config data : " + str(self.configdata))


    def opencode(self, file):
        """
        Open the Operation Code from designated Files
        """
        with open(file) as fin:
            data = json.loads(fin.read())
        return data
    
    def read_data(self, file):
        try:
            self.logger.debug("Excel File "+str(file))
            url_list = []
            work_book = xlrd.open_workbook(file)
            work_sheet = work_book.sheet_by_index(0)
            num_rows = work_sheet.nrows
            num_cols = work_sheet.ncols
            for row in range(1, num_rows):
                chunks = []
                for col in range(0, num_cols):
                    if bool(str(work_sheet.cell_value(row, col)).strip()):
                        chunks.append(work_sheet.cell_value(row, col))
                url_list.append(chunks)
            self.logger.debug(url_list)
            return url_list
        except:
            self.logger.error("Below Exception occurred\n", exc_info=True)
            return []

    def read_data_xlsx(self, file):
        try:
            self.logger.debug("Excel File "+str(file))
            url_list = []
            work_book = openpyxl.load_workbook(file)
            work_sheet = work_book.active
            num_rows = work_sheet.max_row
            num_cols = work_sheet.max_column
            for row in range(2, num_rows+1):
                chunks = []
                if bool(str(work_sheet.cell(row, 1)).strip()):
                        chunks.append(work_sheet.cell(row, 1).value)
                url_list.append(chunks)
            self.logger.debug(url_list)
            return url_list
        except:
            self.logger.error("Below Exception occurred\n", exc_info=True)
            return []    


    def get_LockOwner(self, payload):   
        try:
            self.logger.debug("Checking for page ...")
            _content_root = self.wfoperationdata.get("content root", "/content/pwc")
            _content_root_experience_fragment = self.wfoperationdata.get("content root experience fragment","/content/experience-fragments/pwc")
            if payload.lower().startswith(_content_root) or payload.lower().startswith(_content_root_experience_fragment):
                pageinfo=self.ip+"/libs/wcm/core/content/pageinfo.json?path="+payload
                requestedpage=requests.get(pageinfo,auth=(self.user, self.passwd),timeout=self.timeout)
                if requestedpage.status_code==200:
                    self.logger.debug("Page Found. Getting the lock owner ...")
                    requestedpagejson=requestedpage.json()
                    lockOwner=requestedpagejson["status"]["lockOwner"]
                    if lockOwner:
                        self.logger.debug("Page is locked by: " +str(lockOwner))
                    else:
                        lockOwner="Page Is Not Locked"   
                        self.logger.debug("Page Is Not Locked")
                elif requestedpage.status_code==401:
                    self.logger.debug("Incorrect UserName/Password")  
                    lockOwner="Incorrect UserName/Password"      
                else:
                    self.logger.debug("Page Not Found")  
                    lockOwner="Page Not Found"
            else:
                self.logger.debug("Incorrect Page Path")
                lockOwner="Incorrect Page Path"
            return lockOwner    
        except:
            self.logger.error("Below Exception occurred\n", exc_info=True)
            return 999 
                
    def unlock_process(self, lockOwner, payload):
        try:
            self.logger.debug("Finding Lock Owner User path...")
            extuname=lockOwner
            storeextuname="none"
            extnewpassword=self.commonpasswordforalluser
            if(extuname!=storeextuname):
            ##Get user path
               requesteduserpath=requests.get(self.ip+'/bin/querybuilder.json?path=/home/users&1_property=rep:authorizableId&1_property.value='+extuname,auth=(self.user, self.passwd),timeout=self.timeout)
               val=requesteduserpath.json()
               hits=val["hits"]
               path=[x["path"] for x in hits]
               user_path=path[0]
               post_data={
               "rep:password":extnewpassword
               }
               if requesteduserpath.status_code==200:
                 self.logger.debug("User successfully found in DPE")
                 #Reset password
                 resetpasswordreq=requests.post(self.ip+"/"+user_path,data=post_data,auth=(self.user, self.passwd),timeout=self.timeout)
                 if resetpasswordreq.status_code==200:
                   self.logger.debug("Password reset Successfull")
                 else:
                   self.logger.debug("Password reset cannot be completed. Error") 
                 #End
               else:
                 self.logger.debug("User not found. Error")  
             #End                           
            storeextuname=extuname
            #Unlock Page
            if payload.endswith(".html"):
              payload=payload.replace(".html","")  
            post_data_unlock = {
             "cmd":"unlockPage",
             "_charset_":"utf-8",
             "path":payload
             }
            response = requests.post(self.ip+'/bin/wcmcommand', data=post_data_unlock, auth=(extuname,extnewpassword),timeout=self.timeout)
            time.sleep(self.sleeptime)
            if response.status_code==200:
               self.logger.debug("Page Unlocked Successfully")
               return ["Page Unlocked Successfully",extuname]
            elif response.status_code==401:   
              self.logger.debug("Invalid LockOwner Username/Password")
              return ["Invalid LockOwner Username/Password",extuname]
            elif response.status_code==500:   
               self.logger.debug("Internal Server Error")  
               return ["Internal Server Error",extuname]
            else:
                self.logger.debug("Page Unlock Cannot Be Completed")     
                return "Page Unlock Cannot Be Completed"
            #End
                                 
        except:
            self.logger.error("Below Exception occurred\n", exc_info=True)
            return 999  
        
  

     

    