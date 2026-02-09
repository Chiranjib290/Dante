import requests
from openpyxl import load_workbook
import concurrent.futures
 
name = "cleanup0108.xlsx"
input_file_name=name
 
wb = load_workbook(input_file_name)
ws = wb.active
 
BASE_URL    = 'http://10.195.141.181:4503'
BASE_URL_2    = 'http://10.195.141.180:4503'
USERNAME    = 'admin'
PASSWORD    = 'BsSTG_Mar21'
AUTH    = (USERNAME, PASSWORD)
 
row=1
col=0
 
 
def user_delete(i, path):
    #print(path)
   
    url  = f"{BASE_URL}{path}"
    url_2 = f"{BASE_URL_2}{path}"
    data = {'deleteAuthorizable': ''}
   
    response = requests.get(f"{BASE_URL}{path}.json", auth=AUTH , timeout=5)
    if response.status_code == 200:
        response = requests.post(url, data=data, auth=AUTH, timeout=5)
        if response.status_code == 200:
          output=f"Deleted"
        else:
          output=f"Failed to delete"
          print(output,"1")
    else:
       output=f"Unable to find user"
       print(output,"1")
    ws.cell(row=i, column=2, value=output)  
   
    response_2 = requests.get(f"{BASE_URL_2}{path}.json", auth=AUTH , timeout=5)
    if response_2.status_code == 200:
        response = requests.post(url_2, data=data, auth=AUTH, timeout=5)
        if response.status_code == 200:
          output_2=f"Deleted"
        else:
          output_2=f"Failed to delete"
          print(output_2,"2")
    else:
       output_2=f"Unable to find user"
       print(output_2,"2")
    ws.cell(row=i, column=3, value=output_2)
 
with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
    futures = []
    for i, cell in enumerate(ws['A'][1:], start=2):
        value = cell.value
        futures.append(executor.submit(user_delete, i, value))
 
    concurrent.futures.wait(futures)
 
 
 
output_file_name = name
wb.save(output_file_name)  