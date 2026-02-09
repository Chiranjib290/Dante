import requests
from openpyxl import load_workbook
import concurrent.futures
 
input_file_name="test.xlsx"
wb = load_workbook(input_file_name)
ws = wb.active
 
BASE_URL    = 'http://10.196.83.228:4502'
USERNAME    = 'admin'
PASSWORD    = 'B9HbfFTdV2oq15yotZwyowV4dQUJGC'
AUTH    = (USERNAME, PASSWORD)
 
row=1
col=0
 
 
def user_delete(i, path):
    print(value)
    url  = f"{BASE_URL}{path}"
    data = {'deleteAuthorizable': ''}
    response = requests.post(url, data=data, auth=AUTH, timeout=10)
    if response.status_code == 200:
        output=f"✅Deleted"
    else:
        output=f"❌ Failed to delete"
   
    ws.cell(row=i, column=2, value=output)
 
with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
    futures = []
    for i, cell in enumerate(ws['A'][1:], start=2):
        value = cell.value
        futures.append(executor.submit(user_delete, i, value))
 
    concurrent.futures.wait(futures)
 
 
 
output_file_name = "test.xlsx"
wb.save(output_file_name)  