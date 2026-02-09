from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup

def find_empty_contact_pages(urls):
    """
    Given a list of URLs, return those where the <div class="contactCol-container">
    exists but is empty (no child elements or meaningful text).
    """
    empty_pages = []
    
    for url in urls:
        try:
            response = requests.get(url,auth=("chiranjib.bhattacharyya@in.pwc.com","Change@123456"), timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, "html.parser")
            container = soup.find("div", class_="contactCol-container")
            
            if container:
                # Strip whitespace and check if it has no child elements
                content = container.get_text(strip=True)
                children = container.find_all(recursive=False)
                
                if not content and not children:
                    empty_pages.append(url)
        
        except Exception as e:
            print(f"Error fetching {url}: {e}")
    
    return empty_pages

def list_to_excel(lst, excel_filename):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Result")
    ws.cell(row=1, column=1, value="Payload")
    for idx, data in enumerate(lst, start=2):
        ws.cell(row=idx, column=1, value=data)
    wb.save(excel_filename)

def excel_to_list(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
    return [cell.value for cell in ws['A'][1:]]


if __name__ == "__main__":
    file="tst.xlsx"
    ls = excel_to_list(file)
    rs = find_empty_contact_pages(ls)
    list_to_excel(rs,"rst.xlsx")